"""Business Capability Master Data router — list + Excel import/export."""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any

from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

from app.database import get_db
from app.utils.pagination import PaginationParams, paginated_response
from app.utils.filters import multi_value_int_condition

from app.auth import require_permission, require_role, Role

from app.infrastructure.database.repositories.application_repo import PostgresBizCapabilityRepository
from app.application.application.services import BizCapabilityService

router = APIRouter()

SORT_FIELD_MAP: dict[str, str] = {
    "bcId": "bc_id",
    "parentBcId": "parent_bc_id",
    "bcName": "bc_name",
    "domainL1": "lv1_domain",
    "subDomainL2": "lv2_sub_domain",
    "capabilityGroupL3": "lv3_capability_group",
    "version": "data_version",
}

ALLOWED_SORT_FIELDS = set(SORT_FIELD_MAP.values()) | {
    "bc_id", "parent_bc_id", "bc_name", "lv1_domain", "lv2_sub_domain",
    "lv3_capability_group", "data_version", "level",
}

EXPORT_HEADERS = [
    "Domain(L1)",
    "Sub Domain(L2)",
    "Capability Group(L3)",
    "BC ID",
    "BC Name",
    "BC Name with ID",
    "BC Name CN",
    "Level",
    "Alias",
    "BC BF Description",
    "BG",
    "GEO",
    "BC Biz Owner",
    "BC Biz Team",
    "BC DT Owner",
    "BC DT Team",
]

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "bcid": ("bcid",),
    "parentbcid": ("parentbcid",),
    "bcname": ("bcname",),
    "bcnamewithid": ("bcnamewithid",),
    "bcnamecn": ("bcnamecn",),
    "domainl1": ("domainl1",),
    "subdomainl2": ("subdomainl2",),
    "capabilitygroupl3": ("capabilitygroupl3",),
    "level": ("level",),
    "description": ("description", "bcbfdescription", "bcdescription"),
    "alias": ("alias",),
    "bizgroup": ("bizgroup", "bg"),
    "geo": ("geo",),
    "bizowner": ("bizowner", "bcbizowner"),
    "bizteam": ("bizteam", "bcbizteam"),
    "dtowner": ("dtowner", "bcdtowner"),
    "dtteam": ("dtteam", "bcdtteam"),
    "remark": ("remark",),
    "version": ("version",),
}


def _map_biz_capability(r) -> dict:
    m = r._mapping if hasattr(r, "_mapping") else r
    return {
        "id": int(m["id"]) if m["id"] is not None else None,
        "bcId": m["bc_id"],
        "parentBcId": m["parent_bc_id"],
        "bcName": m["bc_name"],
        "bcNameCn": m["bc_name_cn"],
        "level": m["level"],
        "alias": m["alias"],
        "bcDescription": m["bc_description"],
        "bizGroup": m["biz_group"],
        "geo": m["geo"],
        "bizOwner": m["biz_owner"],
        "bizTeam": m["biz_team"],
        "dtOwner": m["dt_owner"],
        "dtTeam": m["dt_team"],
        "remark": m["remark"],
        "version": m["data_version"],
        "domainL1": m["lv1_domain"],
        "subDomainL2": m["lv2_sub_domain"],
        "capabilityGroupL3": m["lv3_capability_group"],
        "createTime": m["create_time"],
    }


def _build_filters(
    version: str | None,
    domainL1: str | None,
    subDomainL2: str | None,
    capabilityGroupL3: str | None,
    level: str | None,
) -> tuple[str, dict[str, Any]]:
    conditions: list[str] = []
    params: dict[str, Any] = {}

    if version:
        conditions.append("data_version = :version")
        params["version"] = version
    if domainL1:
        conditions.append("lv1_domain = :domainL1")
        params["domainL1"] = domainL1
    if subDomainL2:
        conditions.append("lv2_sub_domain = :subDomainL2")
        params["subDomainL2"] = subDomainL2
    if capabilityGroupL3:
        conditions.append("lv3_capability_group = :capabilityGroupL3")
        params["capabilityGroupL3"] = capabilityGroupL3
    if level:
        conditions.append(multi_value_int_condition("level", "level", level, params))

    return (" AND ".join(conditions) if conditions else "1=1", params)


def _norm_header(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (name or "").strip().lower())


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _resolve_header_indexes(raw_headers: list[str]) -> dict[str, int]:
    normalized_headers = {_norm_header(h): i for i, h in enumerate(raw_headers) if _norm_header(h)}
    resolved: dict[str, int] = {}
    for canonical_key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized_headers:
                resolved[canonical_key] = normalized_headers[alias]
                break
    return resolved


def _extract_biz_capability_label_name(value: str | None, segment_count: int | None = None) -> str | None:
    text = _cell_to_str(value)
    if not text:
        return None

    if segment_count is not None:
        parts = text.split(".")
        if len(parts) > segment_count:
            return ".".join(parts[segment_count:]).strip() or None
        if segment_count == 2 and len(parts) == 2 and re.fullmatch(r"[A-Za-z]+\d+", parts[0]):
            return parts[1].strip() or None

    match = re.match(r"^[A-Za-z0-9]+(?:\.[A-Za-z0-9]+)+\s+(.+)$", text)
    if match:
        return match.group(1).strip() or None
    return text


def _bc_id_prefix(bc_id: str | None, segment_count: int) -> str:
    text = _cell_to_str(bc_id)
    if not text:
        return ""

    parts = text.split(".")
    if len(parts) < segment_count:
        return text
    return ".".join(parts[:segment_count])


def _derive_parent_bc_id(bc_id: str | None) -> str | None:
    text = _cell_to_str(bc_id)
    if not text:
        return None
    if "." not in text:
        return None
    return text.rsplit(".", 1)[0] or None


def _format_biz_capability_hierarchy_label(
    bc_id: str | None,
    label: str | None,
    segment_count: int,
    separator: str = " ",
) -> str | None:
    name = _cell_to_str(label)
    if not name:
        return None

    prefix = _bc_id_prefix(bc_id, segment_count)
    return f"{prefix}{separator}{name}" if prefix else name


def _build_biz_capability_export_row(record: dict[str, Any]) -> list[Any]:
    return [
        _format_biz_capability_hierarchy_label(record.get("bc_id"), record.get("lv1_domain"), 2, "."),
        _format_biz_capability_hierarchy_label(record.get("bc_id"), record.get("lv2_sub_domain"), 3),
        _format_biz_capability_hierarchy_label(record.get("bc_id"), record.get("lv3_capability_group"), 4),
        record.get("bc_id"),
        record.get("bc_name"),
        _format_biz_capability_hierarchy_label(record.get("bc_id"), record.get("bc_name"), 4),
        record.get("bc_name_cn"),
        record.get("level"),
        record.get("alias"),
        record.get("bc_description"),
        record.get("biz_group"),
        record.get("geo"),
        record.get("biz_owner"),
        record.get("biz_team"),
        record.get("dt_owner"),
        record.get("dt_team"),
    ]


def _parse_excel_rows(content: bytes, data_version: str) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Excel support requires openpyxl") from exc

    errors: list[dict[str, str]] = []
    rows_out: list[dict[str, Any]] = []

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], [{"rowNo": "-", "errorResult": "Excel worksheet is missing"}]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], [{"rowNo": "-", "errorResult": "Excel is empty"}]

    raw_headers = [_cell_to_str(h) for h in all_rows[0]]
    idx_map = _resolve_header_indexes(raw_headers)

    required_headers = ["bcid", "level"]
    missing_headers = [h for h in required_headers if h not in idx_map]
    if "bcname" not in idx_map and "bcnamewithid" not in idx_map:
        missing_headers.append("bcname|bcnamewithid")
    if missing_headers:
        return [], [{"rowNo": "-", "errorResult": f"Missing required column(s): {', '.join(missing_headers)}"}]

    def get_val(row: tuple[Any, ...], key: str) -> str:
        idx = idx_map.get(key)
        if idx is None or idx >= len(row):
            return ""
        return _cell_to_str(row[idx])

    seen_bc_ids: set[str] = set()

    for i, row in enumerate(all_rows[1:], start=2):
        if not any(_cell_to_str(c) for c in row):
            continue

        bc_id = get_val(row, "bcid")
        parent_bc_id = get_val(row, "parentbcid") or None
        bc_name = get_val(row, "bcname") or (_extract_biz_capability_label_name(get_val(row, "bcnamewithid"), 4) or "")
        bc_name_cn = get_val(row, "bcnamecn") or None
        domain_l1 = _extract_biz_capability_label_name(get_val(row, "domainl1"), 2)
        sub_domain_l2 = _extract_biz_capability_label_name(get_val(row, "subdomainl2"), 3)
        capability_group_l3 = _extract_biz_capability_label_name(get_val(row, "capabilitygroupl3"), 4)
        level_raw = get_val(row, "level")
        description = get_val(row, "description") or None
        alias = get_val(row, "alias") or None
        biz_group = get_val(row, "bizgroup") or None
        geo = get_val(row, "geo") or None
        biz_owner = get_val(row, "bizowner") or None
        biz_team = get_val(row, "bizteam") or None
        dt_owner = get_val(row, "dtowner") or None
        dt_team = get_val(row, "dtteam") or None
        remark = get_val(row, "remark") or None
        row_version = get_val(row, "version")

        row_errors: list[str] = []
        if not bc_id:
            row_errors.append("BC ID is required")
        if not bc_name:
            row_errors.append("BC Name is required")
        if not level_raw:
            row_errors.append("Level is required")

        parsed_level: int | None = None
        if level_raw:
            try:
                parsed_level = int(float(level_raw))
                if parsed_level < 1 or parsed_level > 4:
                    row_errors.append("Level must be between 1 and 4")
            except Exception:
                row_errors.append("Level must be a number")

        if parsed_level == 1 and not parent_bc_id:
            parent_bc_id = "root"
        elif not parent_bc_id:
            parent_bc_id = _derive_parent_bc_id(bc_id)

        if parsed_level and parsed_level > 1 and not parent_bc_id:
            row_errors.append("Parent BC ID is required for level > 1")

        if bc_id and bc_id in seen_bc_ids:
            row_errors.append("Duplicate BC ID in Excel")
        if bc_id:
            seen_bc_ids.add(bc_id)

        if row_version and row_version != data_version:
            row_errors.append(f"Version mismatch (row={row_version}, input={data_version})")

        if row_errors:
            errors.append({"rowNo": str(i), "errorResult": "; ".join(row_errors)})
            continue

        rows_out.append(
            {
                "bc_id": bc_id,
                "parent_bc_id": parent_bc_id,
                "bc_name": bc_name,
                "bc_name_cn": bc_name_cn,
                "level": parsed_level,
                "alias": alias,
                "bc_description": description,
                "biz_group": biz_group,
                "geo": geo,
                "biz_owner": biz_owner,
                "biz_team": biz_team,
                "dt_owner": dt_owner,
                "dt_team": dt_team,
                "remark": remark,
                "data_version": data_version,
                "lv1_domain": domain_l1,
                "lv2_sub_domain": sub_domain_l2,
                "lv3_capability_group": capability_group_l3,
            }
        )

    if rows_out:
        all_ids = {r["bc_id"] for r in rows_out}
        for idx, r in enumerate(rows_out, start=1):
            parent_id = r.get("parent_bc_id")
            # Allow 'root' as a special value that doesn't need to exist in the file
            if parent_id and parent_id.lower() != "root" and parent_id not in all_ids:
                errors.append(
                    {
                        "rowNo": str(idx + 1),
                        "errorResult": f"Parent BC ID '{parent_id}' does not exist in uploaded file",
                    }
                )

    return rows_out, errors


@router.get("/versions", dependencies=[Depends(require_permission("biz_capability", "read"))])
async def list_biz_capability_versions(db: AsyncSession = Depends(get_db)):
    try:
        query = text("SELECT DISTINCT data_version FROM eam.bcpf_master_data WHERE data_version IS NOT NULL ORDER BY data_version DESC")
        result = await db.execute(query)
        versions = [r.get("version") or r.get("data_version") or r[0] for r in result.mappings().all()]
        return {"versions": versions}
    except Exception as e:
        print(f"Error fetching BizCapability versions: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch BizCapability versions")


@router.get("/filter-options", dependencies=[Depends(require_permission("biz_capability", "read"))])
async def get_biz_capability_filter_options(
    db: AsyncSession = Depends(get_db),
    version: str | None = Query(None),
    domainL1: str | None = Query(None),
    subDomainL2: str | None = Query(None),
):
    try:
        conditions = []
        params = {}
        if version:
            conditions.append("data_version = :version")
            params["version"] = version
        
        # 1. Get Domain L1 list
        l1_query_text = "SELECT DISTINCT lv1_domain FROM eam.bcpf_master_data WHERE level=1 AND lv1_domain IS NOT NULL"
        if conditions:
            l1_query_text += " AND " + " AND ".join(conditions)
        l1_query_text += " ORDER BY lv1_domain ASC"
        
        l1_res = await db.execute(text(l1_query_text), params)
        domains = [r.get("bc_id") or r.get("lv1_domain") or r[0] for r in l1_res.mappings().all()]

        # 2. Get Sub Domain L2 list (filtered by Domain L1 if provided)
        l2_conditions = list(conditions)
        l2_params = dict(params)
        if domainL1:
            l2_conditions.append("lv1_domain = :l1")
            l2_params["l1"] = domainL1
        
        l2_query_text = "SELECT DISTINCT lv2_sub_domain FROM eam.bcpf_master_data WHERE level=2 AND lv2_sub_domain IS NOT NULL"
        if l2_conditions:
            l2_query_text += " AND " + " AND ".join(l2_conditions)
        l2_query_text += " ORDER BY lv2_sub_domain ASC"
        
        l2_res = await db.execute(text(l2_query_text), l2_params)
        sub_domains = [
            r.get("process_name") or r.get("lv2_sub_domain") or r.get("process_id") or r[0]
            for r in l2_res.mappings().all()
        ]

        # 3. Get Capability Group L3 list (filtered by L1 and L2 if provided)
        l3_conditions = list(l2_conditions)
        l3_params = dict(l2_params)
        if subDomainL2:
            l3_conditions.append("lv2_sub_domain = :l2")
            l3_params["l2"] = subDomainL2

        l3_query_text = "SELECT DISTINCT lv3_capability_group FROM eam.bcpf_master_data WHERE level=3 AND lv3_capability_group IS NOT NULL"
        if l3_conditions:
            l3_query_text += " AND " + " AND ".join(l3_conditions)
        l3_query_text += " ORDER BY lv3_capability_group ASC"
        
        l3_res = await db.execute(text(l3_query_text), l3_params)
        capability_groups = [r.get("lv3_capability_group") or r[0] for r in l3_res.mappings().all()]

        return {
            "domainL1": domains,
            "subDomainL2": sub_domains,
            "capabilityGroupL3": capability_groups
        }
    except Exception as e:
        print(f"Error fetching BizCapability filter options: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch BizCapability filter options")


@router.get("", dependencies=[Depends(require_permission("biz_capability", "read"))])
async def list_biz_capability(
    pagination: PaginationParams = Depends(),
    version: str | None = Query(None),
    domainL1: str | None = Query(None),
    subDomainL2: str | None = Query(None),
    capabilityGroupL3: str | None = Query(None),
    level: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
):
    try:
        repo = PostgresBizCapabilityRepository(db)
        service = BizCapabilityService(repo)

        db_sort_field = SORT_FIELD_MAP.get(pagination.sort_field, pagination.sort_field) if pagination.sort_field else "bc_id"
        if db_sort_field not in ALLOWED_SORT_FIELDS:
            db_sort_field = "bc_id"
        sort_order = "ASC" if (not pagination.sort_order) or pagination.sort_order.lower() == "asc" else "DESC"

        rows, total = await service.list_filtered(
            page=pagination.page,
            page_size=pagination.page_size,
            version=version,
            domainL1=domainL1,
            subDomainL2=subDomainL2,
            capabilityGroupL3=capabilityGroupL3,
            level=level,
            sort_field=db_sort_field,
            sort_order=sort_order,
        )

        return paginated_response([_map_biz_capability(r) for r in rows], total, pagination.page, pagination.page_size)
    except Exception as e:
        print(f"Error fetching Business Capability Master Data: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch Business Capability Master Data")


@router.post("/import/validate", dependencies=[Depends(require_role(Role.EA_ADMIN))])
async def validate_biz_capability_import(
    file: UploadFile = File(...),
    dataVersion: str = Form("test"),
):
    data_version = (dataVersion or "test").strip() or "test"

    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        raise HTTPException(status_code=422, detail="CSV validation is not supported; upload Excel or use the legacy importer")
    if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".xlsm")):
        raise HTTPException(status_code=422, detail="Only .xlsx, .xls, and .xlsm files are supported")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10 MB)")

    parsed, errors = _parse_excel_rows(content, data_version)
    return {
        "valid": len(errors) == 0,
        "totalRows": len(parsed) + len({e['rowNo'] for e in errors if e.get('rowNo') and e['rowNo'] != '-'}) ,
        "validRows": len(parsed),
        "errorRows": errors,
    }


@router.post("/import", dependencies=[Depends(require_role(Role.EA_ADMIN))])
async def import_biz_capability_excel(
    db: AsyncSession = Depends(get_db),
    file: UploadFile = File(...),
    dataVersion: str = Form("test"),
):
    data_version = (dataVersion or "test").strip() or "test"

    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        raise HTTPException(status_code=422, detail="CSV import is not supported; convert the file to Excel first")
    if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".xlsm")):
        raise HTTPException(status_code=422, detail="Only .xlsx, .xls, and .xlsm files are supported")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10 MB)")

    parsed, errors = _parse_excel_rows(content, data_version)
    if errors:
        return {
            "success": False,
            "imported": 0,
            "errorRows": errors,
        }

    if not parsed:
        return {
            "success": False,
            "imported": 0,
            "errorRows": [{"rowNo": "-", "errorResult": "No valid data rows found"}],
        }

    try:
        await db.execute(
            text("DELETE FROM eam.bcpf_master_data WHERE data_version = :version"),
            {"version": data_version},
        )

        insert_sql = text(
            """
            INSERT INTO eam.bcpf_master_data (
                bc_id, parent_bc_id, bc_name, bc_name_cn, level,
                alias, bc_description, biz_group, geo, biz_owner,
                biz_team, dt_owner, dt_team, remark, data_version,
                lv1_domain, lv2_sub_domain, lv3_capability_group, create_time
            ) VALUES (
                :bc_id, :parent_bc_id, :bc_name, :bc_name_cn, :level,
                :alias, :bc_description, :biz_group, :geo, :biz_owner,
                :biz_team, :dt_owner, :dt_team, :remark, :data_version,
                :lv1_domain, :lv2_sub_domain, :lv3_capability_group, NOW()
            )
            """
        )
        for row in parsed:
            await db.execute(insert_sql, row)

        await db.commit()
        return {
            "success": True,
            "imported": len(parsed),
            "errorRows": [],
        }
    except Exception as e:
        await db.rollback()
        print(f"Error importing BizCapability Excel: {e}")
        raise HTTPException(status_code=500, detail="Failed to import BizCapability Excel")


@router.get("/export", dependencies=[Depends(require_permission("biz_capability", "read"))])
async def export_biz_capability_excel(
    db: AsyncSession = Depends(get_db),
    version: str | None = Query(None),
    domainL1: str | None = Query(None),
    subDomainL2: str | None = Query(None),
    capabilityGroupL3: str | None = Query(None),
    level: str | None = Query(None),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Excel export requires openpyxl") from exc

    where_clause, params = _build_filters(version, domainL1, subDomainL2, capabilityGroupL3, level)

    query = text(
        f"""
        SELECT
            bc_id,
            parent_bc_id,
            bc_name,
            bc_name_cn,
            lv1_domain,
            lv2_sub_domain,
            lv3_capability_group,
            level,
            bc_description,
            alias,
            biz_group,
            geo,
            biz_owner,
            biz_team,
            dt_owner,
            dt_team,
            remark,
            data_version
        FROM eam.bcpf_master_data
        WHERE {where_clause}
        ORDER BY bc_id ASC
        """
    )

    result = await db.execute(query, params)
    rows = result.mappings().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=500, detail="Failed to create workbook sheet")
    ws.title = "Business Capability Master Data"
    ws.append(EXPORT_HEADERS)

    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    for col_idx in range(1, len(EXPORT_HEADERS) + 1):
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.fill = header_fill
        header_cell.font = header_font
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    for r in rows:
        m = r if isinstance(r, dict) else r
        ws.append(_build_biz_capability_export_row(m))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"biz-capability-master-data-{ts}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
